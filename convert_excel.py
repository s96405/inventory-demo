# convert_excel.py
# 功能：
# 1. 讀取 Excel 多個工作表
# 2. 抓出每個分頁的表頭
# 3. 抓出每筆製令的製程數量
# 4. 輸出 data/inventory_demo.json 給前端使用

import json
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).parent

EXCEL_FILE = BASE_DIR / "excel" / "30-4118-018-XXX.xlsm"
OUTPUT_FILE = BASE_DIR / "data" / "inventory_demo.json"

# 要轉出的 Excel 分頁
TARGET_SHEETS = [
    "018#1",
    "4XX",
    "5XX",
    "6XX",
    "370",
    "400",
    "420",
    "450",
    "470",
    "500",
    "520",
    "550",
    "570",
    "600",
    "620",
    "650",
]

COLUMNS = [
    {"key": "date", "label": "來料日期", "keywords": ["來料日期"], "type": "text"},
    {
        "key": "work_order",
        "label": "製令單號",
        "keywords": ["製令單號"],
        "type": "text",
    },
    {"key": "input_qty", "label": "入料數", "keywords": ["入料數"], "type": "number"},
    {
        "key": "incoming_wait_qc",
        "label": "進料待檢",
        "keywords": ["進料待檢"],
        "type": "number",
    },
    {
        "key": "wait_issue_qty",
        "label": "#01待發料數量",
        "keywords": ["#01待發料數量"],
        "type": "number",
    },
    {
        "key": "process_1_1_qty",
        "label": "#1-1未加工",
        "keywords": ["#1-1未加工"],
        "type": "number",
    },
    {
        "key": "process_1_1_ng",
        "label": "#1-1 NG",
        "keywords": ["#1-1NG", "#1-1 NG"],
        "type": "number",
    },
    {
        "key": "process_1_2_qty",
        "label": "#1-2未加工",
        "keywords": ["#1-2未加工"],
        "type": "number",
    },
    {
        "key": "process_1_2_ng",
        "label": "#1-2 NG",
        "keywords": ["#1-2NG", "#1-2 NG"],
        "type": "number",
    },
    {
        "key": "process_3_qty",
        "label": "#3未加工",
        "keywords": ["#3未加工", "#03未加工"],
        "type": "number",
    },
    {
        "key": "process_3_ng",
        "label": "#3 NG",
        "keywords": ["#3NG", "#3 NG", "#03NG", "#03 NG"],
        "type": "number",
    },
    {
        "key": "process_4_qty",
        "label": "#4未加工",
        "keywords": ["#4未加工", "#04未加工"],
        "type": "number",
    },
    {
        "key": "process_4_ng",
        "label": "#4 NG",
        "keywords": ["#4NG", "#4 NG", "#04NG", "#04 NG"],
        "type": "number",
    },
    {
        "key": "process_5_qty",
        "label": "#5未加工",
        "keywords": ["#5未加工", "#05未加工"],
        "type": "number",
    },
    {
        "key": "process_5_ng",
        "label": "#5 NG",
        "keywords": ["#5NG", "#5 NG", "#05NG", "#05 NG"],
        "type": "number",
    },
    {
        "key": "process_6_qty",
        "label": "#6未加工",
        "keywords": ["#6未加工", "#06未加工"],
        "type": "number",
    },
    {
        "key": "process_6_ng",
        "label": "#6 NG",
        "keywords": ["#6NG", "#6 NG", "#06NG", "#06 NG"],
        "type": "number",
    },
    {
        "key": "process_7_qty",
        "label": "#7未加工",
        "keywords": ["#7未加工", "#07未加工"],
        "type": "number",
    },
    {
        "key": "process_7_ng",
        "label": "#7 NG",
        "keywords": ["#7NG", "#7 NG", "#07NG", "#07 NG"],
        "type": "number",
    },
    {
        "key": "wait_qc_01",
        "label": "#01待檢驗",
        "keywords": ["#01待檢驗"],
        "type": "number",
    },
    {
        "key": "ng_01",
        "label": "#01NG",
        "keywords": ["#01NG", "#01 NG"],
        "type": "number",
    },
    {
        "key": "wait_outsource_qty",
        "label": "#02待委外數量",
        "keywords": ["#02待委外數量"],
        "type": "number",
    },
    {
        "key": "vendor",
        "label": "#02廠商",
        "keywords": ["#02廠商", "廠商"],
        "type": "text",
    },
    {
        "key": "outsource_wait_return_qty",
        "label": "#02待回數量",
        "keywords": ["#02待回數量", "待回數量"],
        "type": "number",
    },
    {
        "key": "outsource_ng",
        "label": "#02委外NG",
        "keywords": ["#02委外NG", "委外NG"],
        "type": "number",
    },
    {
        "key": "wait_outsource_08_qty",
        "label": "#08待委外數量",
        "keywords": ["#08待委外數量", "#8待委外數量"],
        "type": "number",
    },
    {
        "key": "outsource_08_done_qty",
        "label": "#08已委外數量",
        "keywords": ["#08已委外數量", "#8已委外數量"],
        "type": "number",
    },
    {
        "key": "vendor_08",
        "label": "#08廠商",
        "keywords": ["#08廠商", "#8廠商"],
        "type": "text",
    },
    {
        "key": "outsource_08_wait_return_qty",
        "label": "#08待回數量",
        "keywords": ["#08待回數量", "#8待回數量"],
        "type": "number",
    },
    {
        "key": "outsource_08_ng",
        "label": "#08委外NG",
        "keywords": ["#08委外NG", "#8委外NG"],
        "type": "number",
    },
    {
        "key": "wait_qc_02",
        "label": "#02待檢驗",
        "keywords": ["#02待檢驗"],
        "type": "number",
    },
    {
        "key": "ng_02",
        "label": "#02 NG",
        "keywords": ["#02NG", "#02 NG"],
        "type": "number",
    },
    {
        "key": "wait_qc_06",
        "label": "#06待檢驗數量",
        "keywords": ["#06待檢驗數量", "#6待檢驗數量"],
        "type": "number",
    },
    {
        "key": "finished_wait_qc",
        "label": "成品待檢驗",
        "keywords": ["成品待檢驗"],
        "type": "number",
    },
    {
        "key": "not_stock_in_qty",
        "label": "未入庫數量",
        "keywords": ["未入庫數量", "未入庫"],
        "type": "number",
    },
    {
        "key": "stock_in_qty",
        "label": "入庫數量",
        "keywords": ["入庫數量", "入庫"],
        "type": "number",
    },
    {
        "key": "shipped_qty",
        "label": "已出貨數量",
        "keywords": ["已出貨數量", "已出貨"],
        "type": "number",
    },
    {
        "key": "not_allocated_qty",
        "label": "尚未分料數量",
        "keywords": ["尚未分料數量"],
        "type": "number",
    },
    {"key": "bad_qty", "label": "總不良數", "keywords": ["總不良數"], "type": "number"},
    {"key": "yield_rate", "label": "直通率", "keywords": ["直通率"], "type": "number"},
]

ALLOCATION_COLUMNS = [
    {
        "key": "date",
        "label": "日期",
        "keywords": ["日期"],
        "type": "text",
    },
    {
        "key": "from_work_order",
        "label": "第一製程製令",
        "keywords": ["第一製程製令"],
        "type": "text",
    },
    {
        "key": "allocation_type",
        "label": "分配型號",
        "keywords": ["分配型號"],
        "type": "text",
    },
    {
        "key": "to_work_order",
        "label": "第二製程製令",
        "keywords": ["第二製程製令"],
        "type": "text",
    },
    {
        "key": "allocation_qty",
        "label": "分配數量",
        "keywords": ["分配數量"],
        "type": "number",
    },
]


def safe_text(value):
    if value is None:
        return ""

    return str(value).strip()


def safe_number(value):
    if value is None or value == "":
        return 0

    if isinstance(value, (int, float)):
        return value

    text = str(value).replace(",", "").replace("%", "").strip()

    try:
        return float(text)
    except ValueError:
        return 0


def normalize_text(value):
    return safe_text(value).replace(" ", "").replace("\n", "").replace("\r", "")


def find_header_row(sheet):
    for row_index in range(1, min(sheet.max_row, 40) + 1):
        row_values = [normalize_text(cell.value) for cell in sheet[row_index]]
        row_text = "".join(row_values)

        # 最基本一定要有製令單號
        if "製令單號" not in row_text:
            continue

        # 再看有沒有任一種數量欄位
        has_qty_column = (
            "未加工" in row_text
            or "待回數量" in row_text
            or "未入庫" in row_text
            or "入料數" in row_text
            or "總不良數" in row_text
        )

        if has_qty_column:
            return row_index

    return None


def build_column_map(sheet, header_row):
    column_map = {}

    for cell in sheet[header_row]:
        header_text = normalize_text(cell.value)

        if not header_text:
            continue

        for col_def in COLUMNS:
            key = col_def["key"]

            if key in column_map:
                continue

            for keyword in col_def["keywords"]:
                if normalize_text(keyword) in header_text:
                    column_map[key] = cell.column
                    break

    return column_map


def get_cell_value(sheet, row_index, column_map, col_def):
    col_index = column_map.get(col_def["key"])

    if not col_index:
        return "" if col_def["type"] == "text" else 0

    value = sheet.cell(row=row_index, column=col_index).value

    if col_def["type"] == "text":
        return safe_text(value)

    return safe_number(value)


def build_rows(sheet, header_row, column_map):
    rows = []

    for row_index in range(header_row + 1, sheet.max_row + 1):
        work_order_col = column_map.get("work_order")

        if not work_order_col:
            continue

        work_order = safe_text(sheet.cell(row=row_index, column=work_order_col).value)

        if not work_order:
            continue

        row_data = {}

        for col_def in COLUMNS:
            row_data[col_def["key"]] = get_cell_value(
                sheet, row_index, column_map, col_def
            )

        rows.append(row_data)

    return rows


def find_allocation_header_row(sheet):
    """
    找分配明細表頭列。
    """

    for row_index in range(1, min(sheet.max_row, 80) + 1):
        row_values = [normalize_text(cell.value) for cell in sheet[row_index]]
        row_text = "".join(row_values)

        if (
            "第一製程製令" in row_text
            and "分配型號" in row_text
            and "第二製程製令" in row_text
            and "分配數量" in row_text
        ):
            return row_index

    return None


def build_allocation_column_map(sheet, header_row):
    column_map = {}

    for cell in sheet[header_row]:
        header_text = normalize_text(cell.value)

        if not header_text:
            continue

        for col_def in ALLOCATION_COLUMNS:
            key = col_def["key"]

            if key in column_map:
                continue

            for keyword in col_def["keywords"]:
                if normalize_text(keyword) in header_text:
                    column_map[key] = cell.column
                    break

    return column_map


def get_allocation_cell_value(sheet, row_index, column_map, col_def):
    col_index = column_map.get(col_def["key"])

    if not col_index:
        return "" if col_def["type"] == "text" else 0

    value = sheet.cell(row=row_index, column=col_index).value

    if col_def["type"] == "text":
        return safe_text(value)

    return safe_number(value)


def build_allocation_rows(sheet, header_row, column_map):
    rows = []

    for row_index in range(header_row + 1, sheet.max_row + 1):
        from_order_col = column_map.get("from_work_order")
        to_order_col = column_map.get("to_work_order")
        qty_col = column_map.get("allocation_qty")

        from_order = ""
        to_order = ""
        qty = 0

        if from_order_col:
            from_order = safe_text(
                sheet.cell(row=row_index, column=from_order_col).value
            )

        if to_order_col:
            to_order = safe_text(sheet.cell(row=row_index, column=to_order_col).value)

        if qty_col:
            qty = safe_number(sheet.cell(row=row_index, column=qty_col).value)

        if not from_order and not to_order and qty == 0:
            continue

        row_data = {}

        for col_def in ALLOCATION_COLUMNS:
            row_data[col_def["key"]] = get_allocation_cell_value(
                sheet, row_index, column_map, col_def
            )

        rows.append(row_data)

    return rows


def convert_allocation_rows(workbook):
    """
    從所有工作表中尋找分配明細。
    """

    all_rows = []

    print("開始掃描分配明細...")
    print("-" * 30)

    for sheet_name in workbook.sheetnames:
        print(f"正在掃描分配明細工作表：{sheet_name}")

        sheet = workbook[sheet_name]

        header_row = find_allocation_header_row(sheet)

        if not header_row:
            print(f"沒有找到分配明細表頭：{sheet_name}")
            continue

        column_map = build_allocation_column_map(sheet, header_row)
        rows = build_allocation_rows(sheet, header_row, column_map)

        print(f"找到分配明細工作表：{sheet_name}")
        print(f"分配明細表頭列：{header_row}")
        print(f"分配明細筆數：{len(rows)}")
        print("-" * 30)

        all_rows.extend(rows)

    print(f"分配明細掃描完成，總筆數：{len(all_rows)}")
    print("-" * 30)

    return all_rows


def convert_allocation_summary(workbook):
    """
    抓右側分配摘要：
    018-4XX / 018-5XX / 018-6XX
    分配數量、製程中數量、已做好待檢驗數量
    """

    summary_rows = []

    print("開始掃描分配摘要...")
    print("-" * 30)

    for sheet_name in workbook.sheetnames:
        print(f"正在掃描分配摘要工作表：{sheet_name}")

        sheet = workbook[sheet_name]

        for row_index in range(1, min(sheet.max_row, 80) + 1):
            row_values = [normalize_text(cell.value) for cell in sheet[row_index]]
            row_text = "".join(row_values)

            # 找到 018-4XX / 018-5XX / 018-6XX 那一列
            if not (
                "018-4XX" in row_text
                and "018-5XX" in row_text
                and "018-6XX" in row_text
            ):
                continue

            type_row = row_index
            label_row = row_index + 1
            total_row = row_index + 2
            process_row = row_index + 3
            wait_qc_row = row_index + 4

            type_columns = []

            for cell in sheet[type_row]:
                cell_text = normalize_text(cell.value)

                if cell_text in ["018-4XX", "018-5XX", "018-6XX"]:
                    type_columns.append({"type": cell_text, "column": cell.column})

            for item in type_columns:
                col = item["column"]

                summary_rows.append(
                    {
                        "sheet_name": sheet_name,
                        "allocation_type": item["type"],
                        "total_qty": safe_number(
                            sheet.cell(row=total_row, column=col).value
                        ),
                        "in_process_qty": safe_number(
                            sheet.cell(row=process_row, column=col).value
                        ),
                        "waiting_qc_qty": safe_number(
                            sheet.cell(row=wait_qc_row, column=col).value
                        ),
                    }
                )

            print(f"找到分配摘要工作表：{sheet_name}")
            print(f"摘要筆數：{len(type_columns)}")
            print("-" * 30)

            # 找到一次就不用繼續掃同一張
            break

    print(f"分配摘要掃描完成，總筆數：{len(summary_rows)}")
    print("-" * 30)

    return summary_rows


def convert_sheet(workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        print(f"略過，找不到工作表：{sheet_name}")
        return None

    sheet = workbook[sheet_name]

    header_row = find_header_row(sheet)

    if not header_row:
        print(f"略過，找不到表頭列：{sheet_name}")
        return None

    column_map = build_column_map(sheet, header_row)
    rows = build_rows(sheet, header_row, column_map)

    print(f"工作表：{sheet_name}")
    print(f"表頭列：{header_row}")
    print(f"欄位數：{len(column_map)}")
    print(f"資料筆數：{len(rows)}")
    print("-" * 30)

    return {
        "sheet_name": sheet_name,
        "header_row": header_row,
        "column_map": column_map,
        "rows": rows,
    }


def convert_excel_to_json():
    if not EXCEL_FILE.exists():
        print(f"找不到 Excel 檔案：{EXCEL_FILE}")
        return

    workbook = load_workbook(EXCEL_FILE, data_only=True, read_only=True)

    sheets_data = {}

    for sheet_name in TARGET_SHEETS:
        sheet_data = convert_sheet(workbook, sheet_name)

        if sheet_data:
            sheets_data[sheet_name] = sheet_data

    allocation_rows = convert_allocation_rows(workbook)
    allocation_summary = convert_allocation_summary(workbook)

    output_data = {
        "source_file": EXCEL_FILE.name,
        "sheets": sheets_data,
        "allocations": allocation_rows,
        "allocation_summary": allocation_summary,
    }

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    with open(OUTPUT_FILE, "w", encoding="utf-8") as file:
        json.dump(output_data, file, ensure_ascii=False, indent=2)

    print(f"輸出完成：{OUTPUT_FILE}")
    print(f"成功轉出分頁數：{len(sheets_data)}")


if __name__ == "__main__":
    convert_excel_to_json()
